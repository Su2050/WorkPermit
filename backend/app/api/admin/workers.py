"""
人员管理API
"""
import uuid
from typing import Optional, List
from datetime import datetime

from fastapi import APIRouter, Depends, UploadFile, File
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.models import Worker, SysUser, Contractor
from app.api.admin.auth import get_current_user
from app.middleware.tenant import get_tenant_context, TenantQueryFilter
from app.utils.response import success_response, error_response, ErrorCode
from app.utils.pagination import PaginationParams, paginate

router = APIRouter()


class WorkerQuery(PaginationParams):
    """查询人员参数"""
    keyword: Optional[str] = None
    contractor_id: Optional[uuid.UUID] = None
    status: Optional[str] = None
    is_bound: Optional[bool] = None


@router.get("")
async def list_workers(
    query: WorkerQuery = Depends(),
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """获取人员列表"""
    ctx = get_tenant_context()
    
    stmt = select(Worker).order_by(Worker.name)
    stmt = TenantQueryFilter.apply(stmt, ctx)
    
    if query.keyword:
        stmt = stmt.where(
            (Worker.name.ilike(f"%{query.keyword}%")) |
            (Worker.phone.ilike(f"%{query.keyword}%"))
        )
    
    if query.contractor_id:
        stmt = stmt.where(Worker.contractor_id == query.contractor_id)
    
    if query.status:
        stmt = stmt.where(Worker.status == query.status)
    
    if query.is_bound is not None:
        stmt = stmt.where(Worker.is_bound == query.is_bound)
    
    result = await paginate(db, stmt, query)
    
    # 批量获取施工单位信息
    contractor_ids = {w.contractor_id for w in result.items if w.contractor_id}
    contractors_map = {}
    if contractor_ids:
        from app.models import Contractor
        contractors_result = await db.execute(
            select(Contractor).where(Contractor.contractor_id.in_(contractor_ids))
        )
        for contractor in contractors_result.scalars().all():
            contractors_map[contractor.contractor_id] = contractor.name
    
    items = [
        {
            "worker_id": str(worker.worker_id),
            "name": worker.name,
            "phone": _mask_phone(worker.phone),
            "id_no": _mask_id_no(worker.id_no),  # 列表返回脱敏身份证号
            "job_type": worker.job_type,
            "team_name": worker.team_name,
            "contractor_id": str(worker.contractor_id) if worker.contractor_id else None,
            "contractor_name": contractors_map.get(worker.contractor_id) if worker.contractor_id else None,
            "status": worker.status,
            "is_bound": worker.is_bound
        }
        for worker in result.items
    ]
    
    return success_response({
        "items": items,
        "total": result.total,
        "page": result.page,
        "page_size": result.page_size
    })


@router.get("/options")
async def get_worker_options(
    keyword: Optional[str] = None,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """获取人员选项（用于下拉框）"""
    ctx = get_tenant_context()
    
    # 查询活跃人员，如果status为NULL也包含
    stmt = select(Worker).where(
        (Worker.status == "ACTIVE") |
        (Worker.status.is_(None))
    )
    stmt = TenantQueryFilter.apply(stmt, ctx)
    
    if keyword:
        stmt = stmt.where(
            (Worker.name.ilike(f"%{keyword}%")) |
            (Worker.phone.ilike(f"%{keyword}%"))
        )
    
    # 限制返回数量，避免下拉框数据过多
    stmt = stmt.order_by(Worker.name).limit(100)
    
    result = await db.execute(stmt)
    workers = result.scalars().all()
    
    options = [
        {
            "id": str(worker.worker_id),
            "name": worker.name,
            "worker_id": str(worker.worker_id),
            "id_card_masked": _mask_id_no(worker.id_no) if worker.id_no else "",
            "phone_masked": _mask_phone(worker.phone) if worker.phone else ""
        }
        for worker in workers
    ]
    
    return success_response(options)


@router.get("/{worker_id}")
async def get_worker(
    worker_id: uuid.UUID,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """获取人员详情"""
    result = await db.execute(
        select(Worker).where(Worker.worker_id == worker_id)
    )
    worker = result.scalar_one_or_none()
    
    if not worker:
        return error_response(
            code=ErrorCode.NOT_FOUND,
            message="人员不存在"
        )
    
    # 获取施工单位名称
    contractor_name = None
    if worker.contractor_id:
        contractor_result = await db.execute(
            select(Contractor).where(Contractor.contractor_id == worker.contractor_id)
        )
        contractor = contractor_result.scalar_one_or_none()
        if contractor:
            contractor_name = contractor.name
    
    return success_response({
        "worker_id": str(worker.worker_id),
        "name": worker.name,
        "phone": worker.phone,
        "id_no": worker.id_no,  # 详情页返回完整身份证号
        "job_type": worker.job_type,
        "team_name": worker.team_name,
        "contractor_id": str(worker.contractor_id) if worker.contractor_id else None,
        "contractor_name": contractor_name,
        "status": worker.status,
        "is_bound": worker.is_bound,
        "created_at": worker.created_at.isoformat()
    })


@router.post("/sync-from-realname")
async def sync_from_realname(
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    从实名制系统同步人员
    
    拉取实名制系统的人员列表，同步到本地数据库
    """
    from app.adapters.realname_adapter import RealnameAdapter
    
    ctx = get_tenant_context()
    site_id = ctx.accessible_sites[0] if ctx.accessible_sites else None
    
    if not site_id:
        return error_response(
            code=ErrorCode.VALIDATION_ERROR,
            message="无法确定工地"
        )
    
    adapter = RealnameAdapter()
    
    # 获取实名制系统人员
    page = 1
    synced_count = 0
    
    while True:
        result = await adapter.search_workers(page=page, page_size=100)
        items = result.get("items", [])
        
        if not items:
            break
        
        for item in items:
            # 检查是否已存在
            existing = await db.execute(
                select(Worker).where(
                    Worker.site_id == site_id,
                    Worker.id_no == item["id_no"]
                )
            )
            worker = existing.scalar_one_or_none()
            
            if worker:
                # 更新
                worker.name = item["name"]
                worker.phone = item["phone"]
                worker.job_type = item.get("job_type")
                worker.team_name = item.get("team_name")
                worker.face_id = item.get("face_id")
            else:
                # 新增
                worker = Worker(
                    site_id=site_id,
                    name=item["name"],
                    id_no=item["id_no"],
                    phone=item["phone"],
                    job_type=item.get("job_type"),
                    team_name=item.get("team_name"),
                    face_id=item.get("face_id"),
                    status="ACTIVE"
                )
                db.add(worker)
                synced_count += 1
        
        if len(items) < 100:
            break
        
        page += 1
    
    await db.commit()
    
    return success_response({
        "synced_count": synced_count
    }, message=f"同步完成，新增 {synced_count} 人")


def _mask_phone(phone: str) -> str:
    """手机号脱敏"""
    if not phone or len(phone) < 7:
        return phone
    return phone[:3] + "****" + phone[-4:]


def _mask_id_no(id_no: str) -> str:
    """身份证号脱敏"""
    if not id_no or len(id_no) < 10:
        return id_no
    return id_no[:6] + "********" + id_no[-4:]


class WorkerCreate(BaseModel):
    """创建人员请求"""
    site_id: uuid.UUID = Field(...)
    name: str = Field(..., max_length=100)
    phone: str = Field(..., max_length=20)
    id_no: str = Field(..., max_length=18)
    job_type: Optional[str] = Field(None, max_length=50)
    team_name: Optional[str] = Field(None, max_length=100)
    contractor_id: Optional[uuid.UUID] = None


class WorkerUpdate(BaseModel):
    """更新人员请求"""
    name: Optional[str] = Field(None, max_length=100)
    phone: Optional[str] = Field(None, max_length=20)
    id_no: Optional[str] = Field(None, max_length=18)
    job_type: Optional[str] = Field(None, max_length=50)
    team_name: Optional[str] = Field(None, max_length=100)
    contractor_id: Optional[uuid.UUID] = None
    status: Optional[str] = None


@router.post("")
async def create_worker(
    request: WorkerCreate,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """创建人员"""
    ctx = get_tenant_context()
    
    # 验证用户是否有权访问该工地
    if not ctx.is_sys_admin and request.site_id not in ctx.accessible_sites:
        return error_response(
            code=ErrorCode.FORBIDDEN,
            message="无权在该工地创建人员"
        )
    
    site_id = request.site_id
    
    # 检查手机号是否已存在
    existing_phone = await db.execute(
        select(Worker).where(
            Worker.site_id == site_id,
            Worker.phone == request.phone
        )
    )
    if existing_phone.scalar_one_or_none():
        return error_response(
            code=ErrorCode.VALIDATION_ERROR,
            message="该手机号已存在"
        )
    
    # 检查身份证号是否已存在
    existing_id_no = await db.execute(
        select(Worker).where(
            Worker.site_id == site_id,
            Worker.id_no == request.id_no
        )
    )
    if existing_id_no.scalar_one_or_none():
        return error_response(
            code=ErrorCode.VALIDATION_ERROR,
            message="该身份证号已存在"
        )
    
    worker = Worker(
        site_id=site_id,
        name=request.name,
        phone=request.phone,
        id_no=request.id_no,
        job_type=request.job_type,
        team_name=request.team_name,
        contractor_id=request.contractor_id,
        status="ACTIVE"
    )
    db.add(worker)
    await db.commit()
    
    return success_response({
        "worker_id": str(worker.worker_id)
    }, message="人员创建成功")


@router.patch("/{worker_id}")
async def update_worker(
    worker_id: uuid.UUID,
    request: WorkerUpdate,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """更新人员"""
    result = await db.execute(
        select(Worker).where(Worker.worker_id == worker_id)
    )
    worker = result.scalar_one_or_none()
    
    if not worker:
        return error_response(
            code=ErrorCode.NOT_FOUND,
            message="人员不存在"
        )
    
    # 如果更新手机号，检查是否重复
    if request.phone and request.phone != worker.phone:
        existing_phone = await db.execute(
            select(Worker).where(
                Worker.site_id == worker.site_id,
                Worker.phone == request.phone,
                Worker.worker_id != worker_id
            )
        )
        if existing_phone.scalar_one_or_none():
            return error_response(
                code=ErrorCode.VALIDATION_ERROR,
                message="该手机号已存在"
            )
    
    # 如果更新身份证号，检查是否重复
    if request.id_no and request.id_no != worker.id_no:
        existing_id_no = await db.execute(
            select(Worker).where(
                Worker.site_id == worker.site_id,
                Worker.id_no == request.id_no,
                Worker.worker_id != worker_id
            )
        )
        if existing_id_no.scalar_one_or_none():
            return error_response(
                code=ErrorCode.VALIDATION_ERROR,
                message="该身份证号已存在"
            )
    
    if request.name is not None:
        worker.name = request.name
    if request.phone is not None:
        worker.phone = request.phone
    if request.id_no is not None:
        worker.id_no = request.id_no
    if request.job_type is not None:
        worker.job_type = request.job_type
    if request.team_name is not None:
        worker.team_name = request.team_name
    if request.contractor_id is not None:
        worker.contractor_id = request.contractor_id
    if request.status is not None:
        worker.status = request.status
    
    await db.commit()
    
    return success_response({
        "worker_id": str(worker.worker_id)
    }, message="人员更新成功")


@router.delete("/{worker_id}")
async def delete_worker(
    worker_id: uuid.UUID,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """删除人员（软删除）"""
    result = await db.execute(
        select(Worker).where(Worker.worker_id == worker_id)
    )
    worker = result.scalar_one_or_none()
    
    if not worker:
        return error_response(
            code=ErrorCode.NOT_FOUND,
            message="人员不存在"
        )
    
    worker.status = "INACTIVE"
    await db.commit()
    
    return success_response(message="人员删除成功")


@router.post("/import")
async def import_workers(
    file: UploadFile = File(...),
    contractor_id: Optional[uuid.UUID] = None,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    批量导入人员 (P1-4)
    
    1. 解析Excel/CSV文件
    2. 验证数据格式
    3. 批量创建人员记录
    4. 返回导入结果（成功/失败数量）
    
    Excel格式要求：
    - 第一行为表头
    - 必填列：姓名、手机号、身份证号
    - 可选列：施工单位、工种、备注
    """
    import openpyxl
    import re
    from io import BytesIO
    
    ctx = get_tenant_context()
    
    # 验证文件格式
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        return error_response(
            code=ErrorCode.VALIDATION_ERROR,
            message="不支持的文件格式，请上传 Excel 或 CSV 文件"
        )
    
    try:
        # 读取文件内容
        contents = await file.read()
        
        # 解析 Excel
        if file.filename.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(BytesIO(contents))
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        else:
            # 解析 CSV
            import csv
            from io import StringIO
            csv_content = contents.decode('utf-8-sig')
            reader = csv.reader(StringIO(csv_content))
            rows = list(reader)
        
        if len(rows) < 2:
            return error_response(
                code=ErrorCode.VALIDATION_ERROR,
                message="文件内容为空或格式不正确"
            )
        
        # 解析表头
        headers = [str(h).strip() for h in rows[0]]
        
        # 查找必需列的索引
        name_idx = next((i for i, h in enumerate(headers) if '姓名' in h or 'name' in h.lower()), None)
        phone_idx = next((i for i, h in enumerate(headers) if '手机' in h or 'phone' in h.lower()), None)
        id_no_idx = next((i for i, h in enumerate(headers) if '身份证' in h or 'id' in h.lower()), None)
        contractor_idx = next((i for i, h in enumerate(headers) if '施工单位' in h or 'contractor' in h.lower()), None)
        job_type_idx = next((i for i, h in enumerate(headers) if '工种' in h or 'job' in h.lower()), None)
        remark_idx = next((i for i, h in enumerate(headers) if '备注' in h or 'remark' in h.lower()), None)
        
        if name_idx is None or phone_idx is None or id_no_idx is None:
            return error_response(
                code=ErrorCode.VALIDATION_ERROR,
                message="缺少必需列：姓名、手机号、身份证号"
            )
        
        # 验证手机号格式
        def validate_phone(phone):
            if not phone:
                return False
            phone = str(phone).strip()
            return re.match(r'^1[3-9]\d{9}$', phone) is not None
        
        # 验证身份证号格式
        def validate_id_no(id_no):
            if not id_no:
                return False
            id_no = str(id_no).strip()
            return re.match(r'^\d{17}[\dXx]$', id_no) is not None
        
        # 处理数据
        success_count = 0
        failed_count = 0
        failed_rows = []
        
        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                # 提取数据
                name = str(row[name_idx]).strip() if row[name_idx] else None
                phone = str(row[phone_idx]).strip() if row[phone_idx] else None
                id_no = str(row[id_no_idx]).strip() if row[id_no_idx] else None
                
                # 验证必填字段
                if not name:
                    failed_rows.append({"row": row_idx, "reason": "姓名不能为空"})
                    failed_count += 1
                    continue
                
                if not validate_phone(phone):
                    failed_rows.append({"row": row_idx, "reason": "手机号格式不正确"})
                    failed_count += 1
                    continue
                
                if not validate_id_no(id_no):
                    failed_rows.append({"row": row_idx, "reason": "身份证号格式不正确"})
                    failed_count += 1
                    continue
                
                # 检查是否已存在
                existing = await db.execute(
                    select(Worker).where(
                        (Worker.phone == phone) | (Worker.id_no == id_no)
                    )
                )
                if existing.scalar_one_or_none():
                    failed_rows.append({"row": row_idx, "reason": "手机号或身份证号已存在"})
                    failed_count += 1
                    continue
                
                # 获取施工单位ID
                worker_contractor_id = contractor_id
                if contractor_idx is not None and row[contractor_idx]:
                    contractor_name = str(row[contractor_idx]).strip()
                    contractor_result = await db.execute(
                        select(Contractor).where(Contractor.name == contractor_name)
                    )
                    contractor = contractor_result.scalar_one_or_none()
                    if contractor:
                        worker_contractor_id = contractor.contractor_id
                
                # 创建人员记录
                worker = Worker(
                    site_id=ctx.site_id,
                    contractor_id=worker_contractor_id,
                    name=name,
                    phone=phone,
                    id_no=id_no,
                    job_type=str(row[job_type_idx]).strip() if job_type_idx and row[job_type_idx] else None,
                    remark=str(row[remark_idx]).strip() if remark_idx and row[remark_idx] else None,
                    status="ACTIVE",
                    is_bound=False
                )
                db.add(worker)
                success_count += 1
                
            except Exception as e:
                failed_rows.append({"row": row_idx, "reason": str(e)})
                failed_count += 1
        
        # 提交事务
        await db.commit()
        
        return success_response({
            "success_count": success_count,
            "failed_count": failed_count,
            "failed_rows": failed_rows[:100]  # 最多返回100条失败记录
        }, message=f"导入完成：成功 {success_count} 条，失败 {failed_count} 条")
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"导入人员失败: {e}")
        await db.rollback()
        return error_response(
            code=ErrorCode.UNKNOWN_ERROR,
            message=f"导入失败: {str(e)}"
        )


@router.get("/template")
async def download_import_template(
    current_user: SysUser = Depends(get_current_user)
):
    """
    下载人员导入模板 (P1-4 辅助功能)
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "人员导入模板"
    
    # 设置表头
    headers = ["姓名*", "手机号*", "身份证号*", "施工单位", "工种", "备注"]
    ws.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加示例数据
    example_row = ["张三", "13800138000", "110101199001011234", "XX建筑公司", "电工", "示例数据"]
    ws.append(example_row)
    
    # 添加说明
    ws.append([])
    ws.append(["说明："])
    ws.append(["1. 带*号的列为必填项"])
    ws.append(["2. 手机号格式：11位数字，以1开头"])
    ws.append(["3. 身份证号格式：18位，最后一位可以是X"])
    ws.append(["4. 施工单位：如不填写，需在导入时指定"])
    
    # 调整列宽
    column_widths = [15, 15, 20, 20, 15, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=人员导入模板.xlsx"}
    )

