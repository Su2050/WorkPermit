"""
作业票管理API
"""
import uuid
from datetime import date, time, datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.models import (
    WorkTicket, WorkTicketWorker, WorkTicketArea, WorkTicketVideo,
    DailyTicket, DailyTicketWorker, DailyTicketSnapshot,
    Worker, WorkArea, TrainingVideo, SysUser
)
from app.api.admin.auth import get_current_user
from app.middleware.tenant import get_tenant_context, TenantQueryFilter
from app.utils.response import success_response, error_response, ErrorCode
from app.utils.pagination import PaginationParams, paginate
from app.utils.change_compensator import (
    TicketChanges, TicketChangeValidator, TicketChangeCompensator
)
from app.services.audit_service import AuditService

router = APIRouter()


# 请求/响应模型
class WorkTicketCreate(BaseModel):
    """创建作业票请求"""
    title: str = Field(..., max_length=255)
    description: Optional[str] = None
    contractor_id: uuid.UUID
    start_date: date
    end_date: date
    default_access_start_time: str = "06:00:00"
    default_access_end_time: str = "20:00:00"
    default_training_deadline_time: str = "07:30:00"
    worker_ids: List[uuid.UUID]
    area_ids: List[uuid.UUID]
    video_ids: List[uuid.UUID]
    notify_on_publish: bool = True
    daily_reminder_enabled: bool = True
    remark: Optional[str] = None


class WorkTicketUpdate(BaseModel):
    """更新作业票请求 (P0-5)"""
    remark: Optional[str] = None
    access_start_time: Optional[str] = None
    access_end_time: Optional[str] = None
    training_deadline_time: Optional[str] = None
    add_workers: Optional[List[uuid.UUID]] = None
    remove_workers: Optional[List[uuid.UUID]] = None
    add_areas: Optional[List[uuid.UUID]] = None
    remove_areas: Optional[List[uuid.UUID]] = None
    add_videos: Optional[List[uuid.UUID]] = None
    reason: Optional[str] = None


class WorkTicketQuery(PaginationParams):
    """查询作业票参数"""
    status: Optional[str] = None
    contractor_id: Optional[uuid.UUID] = None
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    keyword: Optional[str] = None


@router.get("")
async def list_tickets(
    query: WorkTicketQuery = Depends(),
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    获取作业票列表
    
    支持分页、筛选
    """
    try:
        ctx = get_tenant_context()
        
        # 构建查询
        stmt = select(WorkTicket).options(
            selectinload(WorkTicket.contractor),
            selectinload(WorkTicket.work_ticket_workers),
            selectinload(WorkTicket.work_ticket_areas),
        ).order_by(WorkTicket.created_at.desc())
        
        # 暂时跳过多租户过滤，先让基本功能工作
        # stmt = TenantQueryFilter.apply(stmt, ctx)
    
        # 状态过滤
        if query.status:
            stmt = stmt.where(WorkTicket.status == query.status)
        
        # 施工单位过滤
        if query.contractor_id:
            stmt = stmt.where(WorkTicket.contractor_id == query.contractor_id)
        
        # 日期范围过滤
        if query.start_date:
            stmt = stmt.where(WorkTicket.end_date >= query.start_date)
        if query.end_date:
            stmt = stmt.where(WorkTicket.start_date <= query.end_date)
        
        # 关键词搜索
        if query.keyword:
            stmt = stmt.where(WorkTicket.title.ilike(f"%{query.keyword}%"))
        
        # 分页
        result = await paginate(db, stmt, query)
        
        # 格式化响应
        items = []
        for ticket in result.items:
            # 获取活跃的人员和区域
            active_workers = [w for w in ticket.work_ticket_workers if w.status == "ACTIVE"]
            active_areas = [a for a in ticket.work_ticket_areas if a.status == "ACTIVE"]
            
            # 获取区域信息（需要单独查询）
            area_details = []
            for area_rel in active_areas:
                area_result = await db.execute(
                    select(WorkArea).where(WorkArea.area_id == area_rel.area_id)
                )
                area = area_result.scalar_one_or_none()
                if area:
                    area_details.append({
                        "area_id": str(area.area_id),
                        "name": area.name
                    })
            
            # 计算已完成培训的人员数（简化版：暂时设为0，后续可以优化）
            completed_workers = 0
            total_workers = len(active_workers)
            
            items.append({
                "ticket_id": str(ticket.ticket_id),
                "title": ticket.title,
                "contractor_name": ticket.contractor.name if ticket.contractor else "",
                "contractor": {
                    "contractor_id": str(ticket.contractor.contractor_id),
                    "name": ticket.contractor.name
                } if ticket.contractor else None,
                "areas": area_details,
                "start_date": str(ticket.start_date),
                "end_date": str(ticket.end_date),
                "status": ticket.status,
                "total_workers": total_workers,
                "completed_workers": completed_workers,
                "worker_count": total_workers,
                "area_count": len(active_areas),
                "created_at": ticket.created_at.isoformat()
            })
        
        return success_response({
            "items": items,
            "total": result.total,
            "page": result.page,
            "page_size": result.page_size
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return error_response(
            code=ErrorCode.UNKNOWN_ERROR,
            message=f"获取作业票列表失败: {str(e)}"
        )


@router.post("")
async def create_ticket(
    request: WorkTicketCreate,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    创建作业票
    
    同时生成每日票据和关联数据
    """
    try:
        ctx = get_tenant_context()
        
        # 获取site_id（从施工单位关联）
        from app.models import Contractor
        contractor_result = await db.execute(
            select(Contractor).where(Contractor.contractor_id == request.contractor_id)
        )
        contractor = contractor_result.scalar_one_or_none()
        
        if not contractor:
            return error_response(
                code=ErrorCode.NOT_FOUND,
                message="施工单位不存在"
            )
        
        site_id = contractor.site_id
        
        # 验证日期
        if request.start_date > request.end_date:
            return error_response(
                code=ErrorCode.VALIDATION_ERROR,
                message="开始日期不能晚于结束日期"
            )
        
        # 解析时间
        def parse_time(t: str) -> time:
            parts = t.split(":")
            return time(int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0)
        
        # 创建作业票
        ticket = WorkTicket(
            site_id=site_id,
            contractor_id=request.contractor_id,
            title=request.title,
            description=request.description,
            start_date=request.start_date,
            end_date=request.end_date,
            default_access_start_time=parse_time(request.default_access_start_time),
            default_access_end_time=parse_time(request.default_access_end_time),
            default_training_deadline_time=parse_time(request.default_training_deadline_time),
            notify_on_publish=request.notify_on_publish,
            daily_reminder_enabled=request.daily_reminder_enabled,
            remark=request.remark,
            status="DRAFT",
            created_by=current_user.user_id
        )
        db.add(ticket)
        await db.flush()
        
        now = datetime.now()
        
        # 添加人员关联
        for worker_id in request.worker_ids:
            tw = WorkTicketWorker(
                ticket_id=ticket.ticket_id,
                worker_id=worker_id,
                site_id=site_id,
                added_at=now,
                added_by=current_user.user_id,
                status="ACTIVE"
            )
            db.add(tw)
        
        # 添加区域关联
        for area_id in request.area_ids:
            ta = WorkTicketArea(
                ticket_id=ticket.ticket_id,
                area_id=area_id,
                site_id=site_id,
                added_at=now,
                added_by=current_user.user_id,
                status="ACTIVE"
            )
            db.add(ta)
        
        # 添加视频关联
        for i, video_id in enumerate(request.video_ids):
            tv = WorkTicketVideo(
                ticket_id=ticket.ticket_id,
                video_id=video_id,
                site_id=site_id,
                sequence_order=i + 1,
                added_at=now,
                added_by=current_user.user_id,
                status="ACTIVE"
            )
            db.add(tv)
        
        # 记录审计日志（在commit之前）
        try:
            audit_service = AuditService(db)
            await audit_service.record(
                action="CREATE",
                resource_type="WorkTicket",
                resource_id=ticket.ticket_id,
                resource_name=ticket.title,
                operator_id=current_user.user_id,
                new_value={"title": ticket.title, "status": ticket.status}
            )
        except Exception as e:
            # 审计日志记录失败不影响主流程
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Failed to record audit log: {e}")
        
        await db.commit()
        
        return success_response({
            "ticket_id": str(ticket.ticket_id),
            "status": ticket.status
        }, message="作业票创建成功")
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"创建作业票失败: {e}")
        logger.error(traceback.format_exc())
        await db.rollback()
        return error_response(
            code=ErrorCode.UNKNOWN_ERROR,
            message=f"创建作业票失败: {str(e)}"
        )


@router.get("/{ticket_id}")
async def get_ticket(
    ticket_id: uuid.UUID,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """获取作业票详情"""
    result = await db.execute(
        select(WorkTicket)
        .options(
            selectinload(WorkTicket.contractor),
            selectinload(WorkTicket.work_ticket_workers).selectinload(WorkTicketWorker.worker),
            selectinload(WorkTicket.work_ticket_areas).selectinload(WorkTicketArea.area),
            selectinload(WorkTicket.work_ticket_videos).selectinload(WorkTicketVideo.video),
            selectinload(WorkTicket.daily_tickets),
        )
        .where(WorkTicket.ticket_id == ticket_id)
    )
    ticket = result.scalar_one_or_none()
    
    if not ticket:
        return error_response(
            code=ErrorCode.TICKET_NOT_FOUND,
            message="作业票不存在"
        )
    
    # 构建详情响应
    workers = [
        {
            "worker_id": str(tw.worker.worker_id),
            "name": tw.worker.name,
            "phone": tw.worker.phone,
            "status": tw.status
        }
        for tw in ticket.work_ticket_workers
    ]
    
    areas = [
        {
            "area_id": str(ta.area.area_id),
            "name": ta.area.name,
            "status": ta.status
        }
        for ta in ticket.work_ticket_areas
    ]
    
    videos = [
        {
            "video_id": str(tv.video.video_id),
            "title": tv.video.title,
            "duration_sec": tv.video.duration_sec,
            "sequence_order": tv.sequence_order,
            "status": tv.status
        }
        for tv in ticket.work_ticket_videos
    ]
    
    daily_tickets = [
        {
            "daily_ticket_id": str(dt.daily_ticket_id),
            "date": str(dt.date),
            "status": dt.status
        }
        for dt in ticket.daily_tickets
    ]
    
    return success_response({
        "ticket_id": str(ticket.ticket_id),
        "title": ticket.title,
        "description": ticket.description,
        "contractor": {
            "contractor_id": str(ticket.contractor.contractor_id),
            "name": ticket.contractor.name
        } if ticket.contractor else None,
        "start_date": str(ticket.start_date),
        "end_date": str(ticket.end_date),
        "default_access_start_time": str(ticket.default_access_start_time),
        "default_access_end_time": str(ticket.default_access_end_time),
        "default_training_deadline_time": str(ticket.default_training_deadline_time),
        "status": ticket.status,
        "remark": ticket.remark,
        "workers": workers,
        "areas": areas,
        "videos": videos,
        "daily_tickets": daily_tickets,
        "created_at": ticket.created_at.isoformat(),
        "updated_at": ticket.updated_at.isoformat()
    })


@router.patch("/{ticket_id}")
async def update_ticket(
    ticket_id: uuid.UUID,
    request: WorkTicketUpdate,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    更新作业票 (P0-5: 变更补偿)
    
    支持变更人员、区域、时间窗等
    """
    # 获取作业票
    result = await db.execute(
        select(WorkTicket).where(WorkTicket.ticket_id == ticket_id)
    )
    ticket = result.scalar_one_or_none()
    
    if not ticket:
        return error_response(
            code=ErrorCode.TICKET_NOT_FOUND,
            message="作业票不存在"
        )
    
    if ticket.status == "CANCELLED":
        return error_response(
            code=ErrorCode.TICKET_CANCELLED,
            message="作业票已取消，无法修改"
        )
    
    # 构建变更对象
    changes = TicketChanges(
        remark=request.remark,
        access_start_time=request.access_start_time,
        access_end_time=request.access_end_time,
        training_deadline_time=request.training_deadline_time,
        add_workers=request.add_workers or [],
        remove_workers=request.remove_workers or [],
        add_areas=request.add_areas or [],
        remove_areas=request.remove_areas or [],
        add_videos=request.add_videos or [],
        reason=request.reason
    )
    
    # 校验变更 (P0-5)
    validator = TicketChangeValidator()
    errors = await validator.validate_change(ticket, changes, db)
    
    if errors:
        return error_response(
            code=ErrorCode.TICKET_CHANGE_FORBIDDEN,
            message="变更校验失败",
            data=[{"field": e.field, "message": e.message, "code": e.code} for e in errors]
        )
    
    # 执行变更和补偿 (P0-5)
    audit_service = AuditService(db)
    from app.services.access_service import AccessService
    access_service = AccessService(db)
    
    compensator = TicketChangeCompensator(db, access_service, audit_service)
    result = await compensator.execute_change(ticket, changes, current_user.user_id)
    
    # 更新简单字段
    if request.remark is not None:
        ticket.remark = request.remark
    
    await db.commit()
    
    return success_response(
        data=result,
        message="作业票更新成功"
    )


@router.post("/{ticket_id}/publish")
async def publish_ticket(
    ticket_id: uuid.UUID,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    发布作业票
    
    生成每日票据，发送通知
    """
    result = await db.execute(
        select(WorkTicket)
        .options(
            selectinload(WorkTicket.work_ticket_workers).selectinload(WorkTicketWorker.worker),
            selectinload(WorkTicket.work_ticket_areas).selectinload(WorkTicketArea.area),
            selectinload(WorkTicket.work_ticket_videos).selectinload(WorkTicketVideo.video),
        )
        .where(WorkTicket.ticket_id == ticket_id)
    )
    ticket = result.scalar_one_or_none()
    
    if not ticket:
        return error_response(
            code=ErrorCode.TICKET_NOT_FOUND,
            message="作业票不存在"
        )
    
    if ticket.status != "DRAFT":
        return error_response(
            code=ErrorCode.VALIDATION_ERROR,
            message="只有草稿状态的作业票可以发布"
        )
    
    # 验证关联数据
    active_workers = [w for w in ticket.work_ticket_workers if w.status == "ACTIVE"]
    active_areas = [a for a in ticket.work_ticket_areas if a.status == "ACTIVE"]
    active_videos = [v for v in ticket.work_ticket_videos if v.status == "ACTIVE"]
    
    if not active_workers:
        return error_response(
            code=ErrorCode.VALIDATION_ERROR,
            message="请至少选择一个作业人员"
        )
    
    if not active_areas:
        return error_response(
            code=ErrorCode.VALIDATION_ERROR,
            message="请至少选择一个作业区域"
        )
    
    if not active_videos:
        return error_response(
            code=ErrorCode.VALIDATION_ERROR,
            message="请至少选择一个培训视频"
        )
    
    # 生成每日票据
    from app.services.ticket_service import TicketService
    ticket_service = TicketService(db)
    daily_tickets = await ticket_service.generate_daily_tickets(ticket)
    
    # 更新状态
    ticket.status = "ACTIVE"
    
    await db.commit()
    
    # 发送通知
    if ticket.notify_on_publish:
        from app.tasks.notification import send_notification_task
        for worker in active_workers:
            send_notification_task.delay(
                worker_id=str(worker.worker_id),
                notification_type="TICKET_PUBLISHED",
                daily_ticket_id=None,
                priority=3,
                data={
                    "title": f"新作业票: {ticket.title}",
                    "ticket_title": ticket.title,
                    "start_date": str(ticket.start_date)
                }
            )
    
    return success_response({
        "ticket_id": str(ticket.ticket_id),
        "status": ticket.status,
        "daily_ticket_count": len(daily_tickets)
    }, message="作业票发布成功")


@router.post("/{ticket_id}/cancel")
async def cancel_ticket(
    ticket_id: uuid.UUID,
    reason: str = None,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """取消作业票"""
    result = await db.execute(
        select(WorkTicket).where(WorkTicket.ticket_id == ticket_id)
    )
    ticket = result.scalar_one_or_none()
    
    if not ticket:
        return error_response(
            code=ErrorCode.TICKET_NOT_FOUND,
            message="作业票不存在"
        )
    
    if ticket.status == "CANCELLED":
        return error_response(
            code=ErrorCode.TICKET_CANCELLED,
            message="作业票已取消"
        )
    
    # 更新状态
    ticket.status = "CANCELLED"
    
    # 取消所有相关的日票
    daily_result = await db.execute(
        select(DailyTicket).where(
            DailyTicket.ticket_id == ticket_id,
            DailyTicket.status.in_(["PUBLISHED", "IN_PROGRESS"])
        )
    )
    daily_tickets = daily_result.scalars().all()
    
    for dt in daily_tickets:
        dt.status = "CANCELLED"
        dt.cancel_reason = reason or "作业票取消"
    
    # 撤销所有相关授权
    from app.models import AccessGrant
    from app.services.access_service import AccessService
    
    access_service = AccessService(db)
    
    # 获取所有相关的授权（通过每日票据关联）
    grants_result = await db.execute(
        select(AccessGrant)
        .join(DailyTicket)
        .where(
            DailyTicket.ticket_id == ticket_id,
            AccessGrant.status.in_(["SYNCED", "PENDING_SYNC", "SYNC_FAILED"])
        )
    )
    grants = grants_result.scalars().all()
    
    revoked_count = 0
    for grant in grants:
        success = await access_service.revoke_grant(
            grant.grant_id,
            reason="TICKET_CANCELLED"
        )
        if success:
            revoked_count += 1
    
    # 记录审计日志
    try:
        audit_service = AuditService(db)
        await audit_service.record(
            action="TICKET_CANCEL",
            resource_type="WorkTicket",
            resource_id=ticket.ticket_id,
            resource_name=ticket.title,
            operator_id=current_user.user_id,
            reason=reason or "作业票取消",
            new_value={
                "status": "CANCELLED",
                "revoked_grants_count": revoked_count
            }
        )
    except Exception as e:
        logger.warning(f"Failed to record audit log: {e}")
    
    await db.commit()
    
    return success_response({
        "ticket_id": str(ticket.ticket_id),
        "status": ticket.status,
        "revoked_grants_count": revoked_count
    }, message=f"作业票已取消，已撤销 {revoked_count} 个授权")


@router.post("/{ticket_id}/close")
async def close_ticket(
    ticket_id: uuid.UUID,
    reason: str = None,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    关闭作业票 (P0-1)
    
    关闭作业票，需要撤销所有相关授权
    与取消不同，关闭是正常结束流程
    """
    result = await db.execute(
        select(WorkTicket).where(WorkTicket.ticket_id == ticket_id)
    )
    ticket = result.scalar_one_or_none()
    
    if not ticket:
        return error_response(
            code=ErrorCode.TICKET_NOT_FOUND,
            message="作业票不存在"
        )
    
    if ticket.status == "CANCELLED":
        return error_response(
            code=ErrorCode.TICKET_CANCELLED,
            message="作业票已取消"
        )
    
    if ticket.status == "CLOSED":
        return error_response(
            code=ErrorCode.VALIDATION_ERROR,
            message="作业票已关闭"
        )
    
    # 更新状态为关闭
    ticket.status = "CLOSED"
    
    # 关闭所有相关的日票
    daily_result = await db.execute(
        select(DailyTicket).where(
            DailyTicket.ticket_id == ticket_id,
            DailyTicket.status.in_(["PUBLISHED", "IN_PROGRESS"])
        )
    )
    daily_tickets = daily_result.scalars().all()
    
    for dt in daily_tickets:
        dt.status = "CLOSED"
    
    # 撤销所有相关授权
    from app.models import AccessGrant
    from app.services.access_service import AccessService
    
    access_service = AccessService(db)
    
    # 获取所有相关的授权（通过每日票据关联）
    grants_result = await db.execute(
        select(AccessGrant)
        .join(DailyTicket)
        .where(
            DailyTicket.ticket_id == ticket_id,
            AccessGrant.status.in_(["SYNCED", "PENDING_SYNC", "SYNC_FAILED"])
        )
    )
    grants = grants_result.scalars().all()
    
    revoked_count = 0
    for grant in grants:
        success = await access_service.revoke_grant(
            grant.grant_id,
            reason="TICKET_CLOSED"
        )
        if success:
            revoked_count += 1
    
    # 记录审计日志
    try:
        audit_service = AuditService(db)
        await audit_service.record(
            action="TICKET_CLOSE",
            resource_type="WorkTicket",
            resource_id=ticket.ticket_id,
            resource_name=ticket.title,
            operator_id=current_user.user_id,
            reason=reason or "作业票关闭",
            new_value={
                "status": "CLOSED",
                "revoked_grants_count": revoked_count
            }
        )
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Failed to record audit log: {e}")
    
    await db.commit()
    
    return success_response({
        "ticket_id": str(ticket.ticket_id),
        "status": ticket.status,
        "revoked_grants_count": revoked_count
    }, message=f"作业票已关闭，已撤销 {revoked_count} 个授权")


@router.get("/stats")
async def get_ticket_stats(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    获取作业票统计数据 (P1-1)
    
    返回作业票总数、各状态数量、完成率等统计信息
    """
    from sqlalchemy import func
    
    ctx = get_tenant_context()
    
    # 构建基础查询
    stmt = select(WorkTicket)
    stmt = TenantQueryFilter.apply(stmt, ctx)
    
    # 日期范围过滤
    if start_date:
        stmt = stmt.where(WorkTicket.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        stmt = stmt.where(WorkTicket.created_at <= datetime.combine(end_date, datetime.max.time()))
    
    result = await db.execute(stmt)
    tickets = result.scalars().all()
    
    # 统计各状态数量（WorkTicket 状态: DRAFT/ACTIVE/CANCELLED）
    total_count = len(tickets)
    draft_count = sum(1 for t in tickets if t.status == "DRAFT")
    active_count = sum(1 for t in tickets if t.status == "ACTIVE")
    cancelled_count = sum(1 for t in tickets if t.status == "CANCELLED")
    
    # 计算活跃率
    active_rate = round((active_count / total_count * 100), 2) if total_count > 0 else 0
    
    # 统计人员数量
    total_workers = 0
    for ticket in tickets:
        worker_stmt = select(func.count(WorkTicketWorker.id)).where(
            WorkTicketWorker.ticket_id == ticket.ticket_id,
            WorkTicketWorker.status == "ACTIVE"
        )
        worker_result = await db.execute(worker_stmt)
        total_workers += worker_result.scalar() or 0
    
    # 统计区域数量
    total_areas = 0
    for ticket in tickets:
        area_stmt = select(func.count(WorkTicketArea.id)).where(
            WorkTicketArea.ticket_id == ticket.ticket_id,
            WorkTicketArea.status == "ACTIVE"
        )
        area_result = await db.execute(area_stmt)
        total_areas += area_result.scalar() or 0
    
    return success_response({
        "total_count": total_count,
        "draft_count": draft_count,
        "active_count": active_count,
        "cancelled_count": cancelled_count,
        "active_rate": active_rate,
        "total_workers": total_workers,
        "total_areas": total_areas,
        "start_date": str(start_date) if start_date else None,
        "end_date": str(end_date) if end_date else None
    })


@router.get("/export")
async def export_tickets(
    status: Optional[str] = None,
    contractor_id: Optional[uuid.UUID] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    导出作业票数据为 Excel 格式 (P1-2)
    
    支持按状态、施工单位、日期范围筛选
    """
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from fastapi.responses import StreamingResponse
    
    ctx = get_tenant_context()
    
    # 构建查询
    stmt = select(WorkTicket).options(
        selectinload(WorkTicket.contractor),
        selectinload(WorkTicket.work_ticket_workers),
        selectinload(WorkTicket.work_ticket_areas)
    ).order_by(WorkTicket.created_at.desc())
    
    stmt = TenantQueryFilter.apply(stmt, ctx)
    
    # 筛选条件
    if status:
        stmt = stmt.where(WorkTicket.status == status)
    if contractor_id:
        stmt = stmt.where(WorkTicket.contractor_id == contractor_id)
    if start_date:
        stmt = stmt.where(WorkTicket.end_date >= start_date)
    if end_date:
        stmt = stmt.where(WorkTicket.start_date <= end_date)
    
    result = await db.execute(stmt)
    tickets = result.scalars().all()
    
    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "作业票列表"
    
    # 设置表头
    headers = [
        "作业票ID", "标题", "施工单位", "状态", "开始日期", "结束日期",
        "人员数量", "区域数量", "创建时间", "备注"
    ]
    ws.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 填充数据
    status_map = {
        "DRAFT": "草稿",
        "IN_PROGRESS": "进行中",
        "CLOSED": "已关闭",
        "CANCELLED": "已取消"
    }
    
    for ticket in tickets:
        worker_count = sum(1 for w in ticket.work_ticket_workers if w.status == "ACTIVE")
        area_count = sum(1 for a in ticket.work_ticket_areas if a.status == "ACTIVE")
        
        row = [
            str(ticket.ticket_id),
            ticket.title,
            ticket.contractor.name if ticket.contractor else "",
            status_map.get(ticket.status, ticket.status),
            ticket.start_date.strftime("%Y-%m-%d"),
            ticket.end_date.strftime("%Y-%m-%d"),
            worker_count,
            area_count,
            ticket.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            ticket.remark or ""
        ]
        ws.append(row)
    
    # 调整列宽
    column_widths = [36, 30, 20, 12, 12, 12, 10, 10, 20, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    from datetime import datetime as dt
    filename = f"作业票列表_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/batch-close")
async def batch_close_tickets(
    ticket_ids: list[uuid.UUID],
    reason: str = "批量关闭",
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    批量关闭作业票 (P2-3)
    
    关闭多个作业票，并撤销所有相关授权
    """
    from app.models import AccessGrant
    from app.services.access_service import AccessService
    
    access_service = AccessService(db)
    
    success_count = 0
    failed_count = 0
    total_revoked = 0
    failed_tickets = []
    
    for ticket_id in ticket_ids:
        try:
            # 获取作业票
            result = await db.execute(
                select(WorkTicket).where(WorkTicket.ticket_id == ticket_id)
            )
            ticket = result.scalar_one_or_none()
            
            if not ticket:
                failed_tickets.append({"ticket_id": str(ticket_id), "reason": "作业票不存在"})
                failed_count += 1
                continue
            
            if ticket.status in ["CANCELLED", "CLOSED"]:
                failed_tickets.append({"ticket_id": str(ticket_id), "reason": f"作业票已{ticket.status}"})
                failed_count += 1
                continue
            
            # 更新状态
            ticket.status = "CLOSED"
            
            # 关闭所有相关的日票
            daily_result = await db.execute(
                select(DailyTicket).where(
                    DailyTicket.ticket_id == ticket_id,
                    DailyTicket.status.in_(["PUBLISHED", "IN_PROGRESS"])
                )
            )
            daily_tickets = daily_result.scalars().all()
            
            for dt in daily_tickets:
                dt.status = "CLOSED"
            
            # 撤销所有相关授权
            grants_result = await db.execute(
                select(AccessGrant)
                .join(DailyTicket)
                .where(
                    DailyTicket.ticket_id == ticket_id,
                    AccessGrant.status.in_(["SYNCED", "PENDING_SYNC", "SYNC_FAILED"])
                )
            )
            grants = grants_result.scalars().all()
            
            revoked_count = 0
            for grant in grants:
                success = await access_service.revoke_grant(
                    grant.grant_id,
                    reason="BATCH_TICKET_CLOSED"
                )
                if success:
                    revoked_count += 1
            
            total_revoked += revoked_count
            
            # 记录审计日志
            try:
                audit_service = AuditService(db)
                await audit_service.record(
                    action="BATCH_TICKET_CLOSE",
                    resource_type="WorkTicket",
                    resource_id=ticket.ticket_id,
                    resource_name=ticket.title,
                    operator_id=current_user.user_id,
                    reason=reason,
                    new_value={
                        "status": "CLOSED",
                        "revoked_grants_count": revoked_count
                    }
                )
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Failed to record audit log: {e}")
            
            success_count += 1
            
        except Exception as e:
            failed_tickets.append({"ticket_id": str(ticket_id), "reason": str(e)})
            failed_count += 1
    
    await db.commit()
    
    return success_response({
        "success_count": success_count,
        "failed_count": failed_count,
        "total_revoked_grants": total_revoked,
        "failed_tickets": failed_tickets
    }, message=f"批量关闭完成：成功 {success_count} 个，失败 {failed_count} 个")


@router.post("/batch-cancel")
async def batch_cancel_tickets(
    ticket_ids: list[uuid.UUID],
    reason: str = "批量取消",
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    批量取消作业票 (P2-3)
    
    取消多个作业票，并撤销所有相关授权
    """
    from app.models import AccessGrant
    from app.services.access_service import AccessService
    
    access_service = AccessService(db)
    
    success_count = 0
    failed_count = 0
    total_revoked = 0
    failed_tickets = []
    
    for ticket_id in ticket_ids:
        try:
            # 获取作业票
            result = await db.execute(
                select(WorkTicket).where(WorkTicket.ticket_id == ticket_id)
            )
            ticket = result.scalar_one_or_none()
            
            if not ticket:
                failed_tickets.append({"ticket_id": str(ticket_id), "reason": "作业票不存在"})
                failed_count += 1
                continue
            
            if ticket.status == "CANCELLED":
                failed_tickets.append({"ticket_id": str(ticket_id), "reason": "作业票已取消"})
                failed_count += 1
                continue
            
            # 更新状态
            ticket.status = "CANCELLED"
            
            # 取消所有相关的日票
            daily_result = await db.execute(
                select(DailyTicket).where(
                    DailyTicket.ticket_id == ticket_id,
                    DailyTicket.status.in_(["PUBLISHED", "IN_PROGRESS"])
                )
            )
            daily_tickets = daily_result.scalars().all()
            
            for dt in daily_tickets:
                dt.status = "CANCELLED"
                dt.cancel_reason = reason
            
            # 撤销所有相关授权
            grants_result = await db.execute(
                select(AccessGrant)
                .join(DailyTicket)
                .where(
                    DailyTicket.ticket_id == ticket_id,
                    AccessGrant.status.in_(["SYNCED", "PENDING_SYNC", "SYNC_FAILED"])
                )
            )
            grants = grants_result.scalars().all()
            
            revoked_count = 0
            for grant in grants:
                success = await access_service.revoke_grant(
                    grant.grant_id,
                    reason="BATCH_TICKET_CANCELLED"
                )
                if success:
                    revoked_count += 1
            
            total_revoked += revoked_count
            
            # 记录审计日志
            try:
                audit_service = AuditService(db)
                await audit_service.record(
                    action="BATCH_TICKET_CANCEL",
                    resource_type="WorkTicket",
                    resource_id=ticket.ticket_id,
                    resource_name=ticket.title,
                    operator_id=current_user.user_id,
                    reason=reason,
                    new_value={
                        "status": "CANCELLED",
                        "revoked_grants_count": revoked_count
                    }
                )
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Failed to record audit log: {e}")
            
            success_count += 1
            
        except Exception as e:
            failed_tickets.append({"ticket_id": str(ticket_id), "reason": str(e)})
            failed_count += 1
    
    await db.commit()
    
    return success_response({
        "success_count": success_count,
        "failed_count": failed_count,
        "total_revoked_grants": total_revoked,
        "failed_tickets": failed_tickets
    }, message=f"批量取消完成：成功 {success_count} 个，失败 {failed_count} 个")

