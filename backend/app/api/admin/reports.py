"""
报表统计API
"""
from datetime import date, datetime, timedelta
from typing import Optional
from io import BytesIO
import uuid
import json
import hashlib
import logging

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, cast, Integer
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from app.core.config import settings
from app.core.database import get_db
from app.models import (
    WorkTicket, DailyTicket, DailyTicketWorker, 
    TrainingSession, AccessGrant, AccessEvent, SysUser,
    WorkTicketArea, Worker, WorkArea
)
from app.api.admin.auth import get_current_user
from app.middleware.tenant import get_tenant_context, TenantQueryFilter
from app.utils.response import success_response

router = APIRouter()

logger = logging.getLogger(__name__)

try:
    import redis.asyncio as redis_async  # type: ignore
except Exception:  # pragma: no cover
    redis_async = None

_redis_client = None


async def _get_redis_client():
    global _redis_client
    if redis_async is None:
        return None
    if _redis_client is None:
        _redis_client = redis_async.from_url(
            settings.REDIS_URL, encoding="utf-8", decode_responses=True
        )
    return _redis_client


def _build_dashboard_cache_key(ctx, today: date) -> str:
    role = getattr(ctx, "user_role", None) or "unknown"
    if ctx is None:
        scope = "anonymous"
    elif getattr(ctx, "is_sys_admin", False):
        scope = "all"
    else:
        sites = getattr(ctx, "accessible_sites", None) or []
        if sites:
            raw = ",".join(sorted(str(s) for s in sites))
            scope = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]
        else:
            scope = "none"
    return f"dashboard:v1:{today.isoformat()}:{role}:{scope}"


@router.get("/dashboard")
async def get_dashboard(
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    获取看板数据
    
    返回今日统计、趋势数据
    """
    ctx = get_tenant_context()
    today = date.today()

    # 轻量缓存：避免每次进入Dashboard都打多条聚合SQL（TTL=60s）
    redis_client = await _get_redis_client()
    cache_key = _build_dashboard_cache_key(ctx, today)
    if redis_client is not None:
        try:
            cached = await redis_client.get(cache_key)
            if cached:
                return success_response(json.loads(cached))
        except Exception as e:  # pragma: no cover
            logger.warning(f"Dashboard cache read failed: {e}")
    
    # 今日作业票统计
    stmt = select(func.count(DailyTicket.daily_ticket_id)).where(
        DailyTicket.date == today,
        DailyTicket.status.in_(["PUBLISHED", "IN_PROGRESS"])
    )
    stmt = TenantQueryFilter.apply(stmt, ctx)
    today_tickets_result = await db.execute(stmt)
    today_tickets = today_tickets_result.scalar() or 0
    
    # 今日培训统计
    stmt = select(func.count(DailyTicketWorker.id)).join(DailyTicket).where(
        DailyTicketWorker.training_status == "COMPLETED",
        DailyTicket.date == today
    )
    if ctx is not None and not ctx.is_sys_admin and ctx.accessible_sites:
        stmt = stmt.where(DailyTicket.site_id.in_(ctx.accessible_sites))
    completed_training_result = await db.execute(stmt)
    completed_training = completed_training_result.scalar() or 0
    
    stmt = select(func.count(DailyTicketWorker.id)).join(DailyTicket).where(
        DailyTicket.date == today,
        DailyTicketWorker.status == "ACTIVE"
    )
    if ctx is not None and not ctx.is_sys_admin and ctx.accessible_sites:
        stmt = stmt.where(DailyTicket.site_id.in_(ctx.accessible_sites))
    total_training_result = await db.execute(stmt)
    total_training = total_training_result.scalar() or 0
    
    # 今日授权统计
    stmt = select(func.count(AccessGrant.grant_id)).join(DailyTicket).where(
        DailyTicket.date == today,
        AccessGrant.status == "SYNCED"
    )
    if ctx is not None and not ctx.is_sys_admin and ctx.accessible_sites:
        stmt = stmt.where(DailyTicket.site_id.in_(ctx.accessible_sites))
    synced_grants_result = await db.execute(stmt)
    synced_grants = synced_grants_result.scalar() or 0
    
    stmt = select(func.count(AccessGrant.grant_id)).join(DailyTicket).where(
        DailyTicket.date == today,
        AccessGrant.status.in_(["PENDING_SYNC", "SYNC_FAILED"])
    )
    if ctx is not None and not ctx.is_sys_admin and ctx.accessible_sites:
        stmt = stmt.where(DailyTicket.site_id.in_(ctx.accessible_sites))
    pending_grants_result = await db.execute(stmt)
    pending_grants = pending_grants_result.scalar() or 0
    
    # 今日进出统计
    today_start = datetime.combine(today, datetime.min.time())
    today_end = datetime.combine(today, datetime.max.time())
    
    stmt = select(func.count(AccessEvent.event_id)).where(
        AccessEvent.event_time >= today_start,
        AccessEvent.event_time <= today_end,
        AccessEvent.result == "PASS"
    )
    stmt = TenantQueryFilter.apply(stmt, ctx)
    pass_events_result = await db.execute(stmt)
    pass_events = pass_events_result.scalar() or 0
    
    stmt = select(func.count(AccessEvent.event_id)).where(
        AccessEvent.event_time >= today_start,
        AccessEvent.event_time <= today_end,
        AccessEvent.result == "DENY"
    )
    stmt = TenantQueryFilter.apply(stmt, ctx)
    deny_events_result = await db.execute(stmt)
    deny_events = deny_events_result.scalar() or 0
    
    # 计算同步率
    total_grants = synced_grants + pending_grants
    sync_rate = round(synced_grants / total_grants * 100, 1) if total_grants > 0 else 0
    
    # 统计今日作业状态（按 DailyTicket 状态分组统计）
    stmt = select(
        DailyTicket.status,
        func.count(DailyTicket.daily_ticket_id)
    ).where(
        DailyTicket.date == today
    ).group_by(DailyTicket.status)
    stmt = TenantQueryFilter.apply(stmt, ctx)
    status_result = await db.execute(stmt)
    status_counts = {row[0]: row[1] for row in status_result.all()}
    
    # 映射状态到前端需要的格式
    today_status = {
        "notStarted": status_counts.get("DRAFT", 0) + status_counts.get("PUBLISHED", 0),
        "inProgress": status_counts.get("IN_PROGRESS", 0),
        "completed": status_counts.get("EXPIRED", 0),
        "failed": status_counts.get("CANCELLED", 0)
    }
    
    # 获取进行中的作业票数量（ACTIVE 状态表示已发布正在进行中）
    stmt = select(func.count(WorkTicket.ticket_id)).where(
        WorkTicket.status == "ACTIVE"
    )
    stmt = TenantQueryFilter.apply(stmt, ctx)
    active_tickets_result = await db.execute(stmt)
    active_tickets = active_tickets_result.scalar() or 0
    
    # 获取待处理作业票（已发布且正在进行中的）
    stmt = select(WorkTicket).options(
        selectinload(WorkTicket.contractor)
    ).where(
        WorkTicket.status == "ACTIVE"
    ).order_by(WorkTicket.created_at.desc()).limit(5)
    stmt = TenantQueryFilter.apply(stmt, ctx)
    pending_tickets_result = await db.execute(stmt)
    pending_tickets = pending_tickets_result.scalars().all()

    # 优化：进度统计由 N+1 查询改成一次聚合查询
    progress_map: dict[str, dict] = {}
    ticket_ids = [t.ticket_id for t in pending_tickets]
    if ticket_ids:
        stmt_progress = (
            select(
                DailyTicket.ticket_id,
                func.count(DailyTicketWorker.id),
                func.sum(cast(DailyTicketWorker.completed_video_count, Integer)),
                func.sum(cast(DailyTicketWorker.total_video_count, Integer)),
            )
            .join(DailyTicket, DailyTicketWorker.daily_ticket_id == DailyTicket.daily_ticket_id)
            .where(
                DailyTicket.date == today,
                DailyTicket.ticket_id.in_(ticket_ids),
                DailyTicketWorker.status == "ACTIVE",
            )
            .group_by(DailyTicket.ticket_id)
        )
        if ctx is not None and not ctx.is_sys_admin and ctx.accessible_sites:
            stmt_progress = stmt_progress.where(DailyTicket.site_id.in_(ctx.accessible_sites))
        progress_result = await db.execute(stmt_progress)
        for tid, worker_cnt, completed_sum, total_sum in progress_result.all():
            progress_map[str(tid)] = {
                "workers": worker_cnt or 0,
                "completed": completed_sum or 0,
                "total": total_sum or 0,
            }

    pending_tickets_list = []
    for ticket in pending_tickets:
        prog = progress_map.get(str(ticket.ticket_id), {"workers": 0, "completed": 0, "total": 0})
        total_workers = prog["workers"]
        total_completed = prog["completed"]
        total_videos = prog["total"] or 0
        # 兼容历史数据（若 total_video_count 未填）
        if not total_videos:
            total_videos = total_workers * 3
        progress = round(total_completed / total_videos * 100, 0) if total_videos > 0 else 0
        
        pending_tickets_list.append({
            "id": str(ticket.ticket_id),
            "title": ticket.title,
            "contractor": ticket.contractor.name if ticket.contractor else "",
            "workerCount": total_workers,
            "progress": int(progress),
            "status": ticket.status
        })
    
    payload = {
        "stats": {
            "todayTasks": today_tickets,
            "activeTickets": active_tickets,
            "todayTrainings": completed_training,
            "accessGrants": synced_grants,
            "syncRate": sync_rate
        },
        "todayStatus": today_status,
        "pendingTickets": pending_tickets_list,
        "recentAlerts": []  # 暂时返回空数组，后续可以添加告警数据
    }

    if redis_client is not None:
        try:
            await redis_client.set(cache_key, json.dumps(payload, ensure_ascii=False), ex=60)
        except Exception as e:  # pragma: no cover
            logger.warning(f"Dashboard cache write failed: {e}")

    return success_response(payload)


@router.get("/dashboard/stats")
async def get_dashboard_stats(
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    获取看板首屏统计（轻量）
    - 只返回 stats（用于首屏4张卡片）
    - 更适合强缓存/快速返回
    """
    ctx = get_tenant_context()
    today = date.today()

    redis_client = await _get_redis_client()
    cache_key = _build_dashboard_cache_key(ctx, today) + ":stats"
    if redis_client is not None:
        try:
            cached = await redis_client.get(cache_key)
            if cached:
                return success_response(json.loads(cached))
        except Exception as e:  # pragma: no cover
            logger.warning(f"Dashboard stats cache read failed: {e}")

    # 今日作业票统计（今日日票）
    stmt = select(func.count(DailyTicket.daily_ticket_id)).where(
        DailyTicket.date == today,
        DailyTicket.status.in_(["PUBLISHED", "IN_PROGRESS"])
    )
    stmt = TenantQueryFilter.apply(stmt, ctx)
    today_tickets_result = await db.execute(stmt)
    today_tickets = today_tickets_result.scalar() or 0

    # 今日培训完成数（完成的人员条目数）
    stmt = select(func.count(DailyTicketWorker.id)).join(DailyTicket).where(
        DailyTicketWorker.training_status == "COMPLETED",
        DailyTicket.date == today
    )
    if ctx is not None and not ctx.is_sys_admin and ctx.accessible_sites:
        stmt = stmt.where(DailyTicket.site_id.in_(ctx.accessible_sites))
    completed_training_result = await db.execute(stmt)
    completed_training = completed_training_result.scalar() or 0

    # 今日授权数（已同步）
    stmt = select(func.count(AccessGrant.grant_id)).join(DailyTicket).where(
        DailyTicket.date == today,
        AccessGrant.status == "SYNCED"
    )
    if ctx is not None and not ctx.is_sys_admin and ctx.accessible_sites:
        stmt = stmt.where(DailyTicket.site_id.in_(ctx.accessible_sites))
    synced_grants_result = await db.execute(stmt)
    synced_grants = synced_grants_result.scalar() or 0

    # 计算同步率需要分母（同步+待同步/失败）
    stmt = select(func.count(AccessGrant.grant_id)).join(DailyTicket).where(
        DailyTicket.date == today,
        AccessGrant.status.in_(["PENDING_SYNC", "SYNC_FAILED"])
    )
    if ctx is not None and not ctx.is_sys_admin and ctx.accessible_sites:
        stmt = stmt.where(DailyTicket.site_id.in_(ctx.accessible_sites))
    pending_grants_result = await db.execute(stmt)
    pending_grants = pending_grants_result.scalar() or 0

    total_grants = synced_grants + pending_grants
    sync_rate = round(synced_grants / total_grants * 100, 1) if total_grants > 0 else 0

    # 进行中的作业票数量（ACTIVE 状态表示已发布正在进行中）
    stmt = select(func.count(WorkTicket.ticket_id)).where(
        WorkTicket.status == "ACTIVE"
    )
    stmt = TenantQueryFilter.apply(stmt, ctx)
    active_tickets_result = await db.execute(stmt)
    active_tickets = active_tickets_result.scalar() or 0

    # 统计今日作业状态（按 DailyTicket 状态分组统计）
    stmt = select(
        DailyTicket.status,
        func.count(DailyTicket.daily_ticket_id)
    ).where(
        DailyTicket.date == today
    ).group_by(DailyTicket.status)
    stmt = TenantQueryFilter.apply(stmt, ctx)
    status_result = await db.execute(stmt)
    status_counts = {row[0]: row[1] for row in status_result.all()}
    
    # 映射状态到前端需要的格式
    today_status = {
        "notStarted": status_counts.get("DRAFT", 0) + status_counts.get("PUBLISHED", 0),
        "inProgress": status_counts.get("IN_PROGRESS", 0),
        "completed": status_counts.get("EXPIRED", 0),
        "failed": status_counts.get("CANCELLED", 0)
    }

    payload = {
        "stats": {
            "todayTasks": today_tickets,
            "activeTickets": active_tickets,
            "todayTrainings": completed_training,
            "accessGrants": synced_grants,
            "syncRate": sync_rate
        },
        "todayStatus": today_status
    }

    if redis_client is not None:
        try:
            await redis_client.set(cache_key, json.dumps(payload, ensure_ascii=False), ex=60)
        except Exception as e:  # pragma: no cover
            logger.warning(f"Dashboard stats cache write failed: {e}")

    return success_response(payload)


@router.get("/reconciliation")
async def get_reconciliation_report(
    date_str: Optional[str] = None,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    获取对账报告 (P0-6)
    
    显示同步状态不一致的授权
    """
    ctx = get_tenant_context()
    report_date = date.fromisoformat(date_str) if date_str else date.today()
    
    # 获取卡住的授权（超过10分钟仍未同步）
    cutoff = datetime.now() - timedelta(minutes=10)
    
    stmt = select(AccessGrant).join(DailyTicket).where(
        DailyTicket.date == report_date,
        AccessGrant.status.in_(["PENDING_SYNC", "SYNC_FAILED"]),
        AccessGrant.created_at < cutoff
    )
    stmt = TenantQueryFilter.apply(stmt, ctx)
    
    result = await db.execute(stmt)
    stuck_grants = result.scalars().all()
    
    items = [
        {
            "grant_id": str(grant.grant_id),
            "worker_id": str(grant.worker_id),
            "area_id": str(grant.area_id),
            "status": grant.status,
            "sync_attempt_count": grant.sync_attempt_count,
            "last_sync_at": grant.last_sync_at.isoformat() if grant.last_sync_at else None,
            "sync_error_msg": grant.sync_error_msg
        }
        for grant in stuck_grants
    ]
    
    return success_response({
        "date": str(report_date),
        "total_stuck": len(items),
        "items": items
    })


@router.get("/training-progress")
async def get_training_progress(
    ticket_id: Optional[uuid.UUID] = None,
    date_str: Optional[str] = None,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    获取培训进度报告
    
    按作业票或日期查看培训完成情况
    """
    ctx = get_tenant_context()
    
    query_date = date.fromisoformat(date_str) if date_str else date.today()
    
    stmt = select(DailyTicketWorker, DailyTicket).join(
        DailyTicket, DailyTicketWorker.daily_ticket_id == DailyTicket.daily_ticket_id
    ).where(
        DailyTicket.date == query_date,
        DailyTicketWorker.status == "ACTIVE"
    )
    
    if ticket_id:
        stmt = stmt.where(DailyTicket.ticket_id == ticket_id)
    
    result = await db.execute(stmt)
    rows = result.all()
    
    # 统计各状态数量
    status_counts = {
        "NOT_STARTED": 0,
        "IN_LEARNING": 0,
        "COMPLETED": 0,
        "FAILED": 0
    }
    
    items = []
    for dtw, dt in rows:
        status_counts[dtw.training_status] = status_counts.get(dtw.training_status, 0) + 1
        items.append({
            "daily_ticket_id": str(dtw.daily_ticket_id),
            "worker_id": str(dtw.worker_id),
            "total_videos": dtw.total_video_count,
            "completed_videos": dtw.completed_video_count,
            "training_status": dtw.training_status,
            "authorized": dtw.authorized
        })
    
    total = len(items)
    
    return success_response({
        "date": str(query_date),
        "total": total,
        "status_counts": status_counts,
        "completion_rate": round(status_counts["COMPLETED"] / total * 100, 1) if total > 0 else 0,
        "items": items[:100]  # 限制返回数量
    })


@router.get("/trend")
async def get_trend(
    metric: str = "completion_rate",  # completion_rate, sync_rate, training_duration
    days: int = 7,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    获取趋势数据
    
    metric: 指标类型
    - completion_rate: 培训完成率趋势
    - sync_rate: 同步成功率趋势
    - training_duration: 培训时长趋势
    
    days: 查询天数（默认7天）
    """
    ctx = get_tenant_context()
    
    from datetime import timedelta
    end_date = date.today()
    start_date = end_date - timedelta(days=days)
    
    trend_data = []
    current_date = start_date
    
    while current_date <= end_date:
        if metric == "completion_rate":
            # 培训完成率
            stmt = select(DailyTicketWorker).join(DailyTicket).where(
                DailyTicket.date == current_date,
                DailyTicketWorker.status == "ACTIVE"
            )
            stmt = TenantQueryFilter.apply(stmt, ctx)
            
            result = await db.execute(stmt)
            workers = result.scalars().all()
            
            total = len(workers)
            completed = sum(1 for w in workers if w.training_status == "COMPLETED")
            rate = round((completed / total * 100), 2) if total > 0 else 0
            
            trend_data.append({
                "date": str(current_date),
                "value": rate
            })
            
        elif metric == "sync_rate":
            # 同步成功率
            from app.models import AccessGrant
            from datetime import datetime
            
            stmt = select(AccessGrant).where(
                AccessGrant.created_at >= datetime.combine(current_date, datetime.min.time()),
                AccessGrant.created_at <= datetime.combine(current_date, datetime.max.time())
            )
            stmt = TenantQueryFilter.apply(stmt, ctx)
            
            result = await db.execute(stmt)
            grants = result.scalars().all()
            
            total = len(grants)
            synced = sum(1 for g in grants if g.status == "SYNCED")
            rate = round((synced / total * 100), 2) if total > 0 else 0
            
            trend_data.append({
                "date": str(current_date),
                "value": rate
            })
            
        elif metric == "training_duration":
            # 平均培训时长
            from app.models import TrainingSession
            
            session_stmt = select(
                func.avg(func.extract('epoch', TrainingSession.ended_at - TrainingSession.started_at))
            ).join(DailyTicketWorker).join(DailyTicket).where(
                DailyTicket.date == current_date,
                TrainingSession.status == "COMPLETED"
            )
            session_stmt = TenantQueryFilter.apply(session_stmt, ctx)
            
            session_result = await db.execute(session_stmt)
            avg_seconds = session_result.scalar() or 0
            avg_minutes = round(avg_seconds / 60, 1) if avg_seconds else 0
            
            trend_data.append({
                "date": str(current_date),
                "value": avg_minutes
            })
        
        current_date += timedelta(days=1)
    
    return success_response({
        "metric": metric,
        "days": days,
        "data": trend_data
    })


@router.get("/access-events")
async def get_access_events(
    date_str: Optional[str] = None,
    result_filter: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    获取进出事件记录
    
    支持按日期和结果筛选
    """
    ctx = get_tenant_context()
    
    query_date = date.fromisoformat(date_str) if date_str else date.today()
    date_start = datetime.combine(query_date, datetime.min.time())
    date_end = datetime.combine(query_date, datetime.max.time())
    
    stmt = select(AccessEvent).where(
        AccessEvent.event_time >= date_start,
        AccessEvent.event_time <= date_end
    ).order_by(AccessEvent.event_time.desc())
    
    stmt = TenantQueryFilter.apply(stmt, ctx)
    
    if result_filter:
        stmt = stmt.where(AccessEvent.result == result_filter)
    
    # 分页
    total_stmt = select(func.count()).select_from(stmt.subquery())
    total_result = await db.execute(total_stmt)
    total = total_result.scalar() or 0
    
    stmt = stmt.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(stmt)
    events = result.scalars().all()
    
    items = [
        {
            "event_id": str(event.event_id),
            "worker_id": str(event.worker_id) if event.worker_id else None,
            "device_id": event.device_id,
            "event_time": event.event_time.isoformat(),
            "direction": event.direction,
            "result": event.result,
            "reason_code": event.reason_code,
            "reason_message": event.reason_message
        }
        for event in events
    ]
    
    return success_response({
        "date": str(query_date),
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": items
    })


@router.get("/training-stats")
async def get_training_stats(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    获取培训统计数据 (P0-2)
    
    返回培训完成率、平均培训时长等统计信息
    """
    ctx = get_tenant_context()
    
    # 解析日期范围
    query_start_date = date.fromisoformat(start_date) if start_date else date.today() - timedelta(days=7)
    query_end_date = date.fromisoformat(end_date) if end_date else date.today()
    
    # 获取时间范围内的所有培训数据
    stmt = select(DailyTicketWorker).join(
        DailyTicket, DailyTicketWorker.daily_ticket_id == DailyTicket.daily_ticket_id
    ).where(
        DailyTicket.date >= query_start_date,
        DailyTicket.date <= query_end_date,
        DailyTicketWorker.status == "ACTIVE"
    )
    stmt = TenantQueryFilter.apply(stmt, ctx)
    
    result = await db.execute(stmt)
    workers = result.scalars().all()
    
    # 统计数据
    total_workers = len(workers)
    completed_count = sum(1 for w in workers if w.training_status == "COMPLETED")
    in_learning_count = sum(1 for w in workers if w.training_status == "IN_LEARNING")
    not_started_count = sum(1 for w in workers if w.training_status == "NOT_STARTED")
    failed_count = sum(1 for w in workers if w.training_status == "FAILED")
    
    # 计算完成率
    completion_rate = round((completed_count / total_workers * 100), 2) if total_workers > 0 else 0
    
    # 计算平均培训时长（分钟）
    start_datetime = datetime.combine(query_start_date, datetime.min.time())
    end_datetime = datetime.combine(query_end_date, datetime.max.time())
    
    session_stmt = select(
        func.avg(func.extract('epoch', TrainingSession.ended_at - TrainingSession.started_at))
    ).select_from(TrainingSession).join(
        DailyTicket, TrainingSession.daily_ticket_id == DailyTicket.daily_ticket_id
    ).where(
        DailyTicket.date >= query_start_date,
        DailyTicket.date <= query_end_date,
        TrainingSession.status == "COMPLETED"
    )
    session_stmt = TenantQueryFilter.apply(session_stmt, ctx)
    
    session_result = await db.execute(session_stmt)
    avg_seconds = session_result.scalar() or 0
    avg_duration_minutes = round(avg_seconds / 60, 1) if avg_seconds else 0
    
    # 获取总视频完成数
    total_videos_completed = sum(w.completed_video_count for w in workers if w.completed_video_count)
    total_videos_required = sum(w.total_video_count for w in workers if w.total_video_count)
    
    return success_response({
        "start_date": str(query_start_date),
        "end_date": str(query_end_date),
        "total_workers": total_workers,
        "completed_count": completed_count,
        "in_learning_count": in_learning_count,
        "not_started_count": not_started_count,
        "failed_count": failed_count,
        "completion_rate": completion_rate,
        "avg_duration_minutes": avg_duration_minutes,
        "total_videos_completed": total_videos_completed,
        "total_videos_required": total_videos_required,
        "video_completion_rate": round((total_videos_completed / total_videos_required * 100), 2) if total_videos_required > 0 else 0
    })


@router.get("/access-sync-stats")
async def get_access_sync_stats(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    获取门禁同步统计数据 (P0-3)
    
    返回同步成功率、失败次数等统计信息
    """
    ctx = get_tenant_context()
    
    # 解析日期范围
    query_start_date = date.fromisoformat(start_date) if start_date else date.today() - timedelta(days=7)
    query_end_date = date.fromisoformat(end_date) if end_date else date.today()
    
    start_datetime = datetime.combine(query_start_date, datetime.min.time())
    end_datetime = datetime.combine(query_end_date, datetime.max.time())
    
    # 获取时间范围内的所有授权记录
    stmt = select(AccessGrant).where(
        AccessGrant.created_at >= start_datetime,
        AccessGrant.created_at <= end_datetime
    )
    stmt = TenantQueryFilter.apply(stmt, ctx)
    
    result = await db.execute(stmt)
    grants = result.scalars().all()
    
    # 统计数据
    total_count = len(grants)
    synced_count = sum(1 for g in grants if g.status == "SYNCED")
    pending_count = sum(1 for g in grants if g.status == "PENDING_SYNC")
    failed_count = sum(1 for g in grants if g.status == "SYNC_FAILED")
    revoked_count = sum(1 for g in grants if g.status == "REVOKED")
    
    # 计算同步成功率
    sync_rate = round((synced_count / total_count * 100), 2) if total_count > 0 else 0
    
    # 计算平均同步时长（秒）
    synced_grants = [g for g in grants if g.status == "SYNCED" and g.last_sync_at]
    avg_sync_duration = 0
    if synced_grants:
        total_sync_time = sum(
            (g.last_sync_at - g.created_at).total_seconds() 
            for g in synced_grants 
            if g.last_sync_at and g.created_at
        )
        avg_sync_duration = round(total_sync_time / len(synced_grants), 1)
    
    # 统计失败原因
    failure_reasons = {}
    for grant in grants:
        if grant.status == "SYNC_FAILED" and grant.sync_error_msg:
            reason = grant.sync_error_msg[:50]  # 截取前50个字符
            failure_reasons[reason] = failure_reasons.get(reason, 0) + 1
    
    # 获取重试次数统计
    max_retry_count = max((g.sync_attempt_count for g in grants), default=0)
    avg_retry_count = round(sum(g.sync_attempt_count for g in grants) / total_count, 1) if total_count > 0 else 0
    
    return success_response({
        "start_date": str(query_start_date),
        "end_date": str(query_end_date),
        "total_count": total_count,
        "synced_count": synced_count,
        "pending_count": pending_count,
        "failed_count": failed_count,
        "revoked_count": revoked_count,
        "sync_rate": sync_rate,
        "avg_sync_duration_seconds": avg_sync_duration,
        "max_retry_count": max_retry_count,
        "avg_retry_count": avg_retry_count,
        "failure_reasons": [
            {"reason": reason, "count": count}
            for reason, count in sorted(failure_reasons.items(), key=lambda x: x[1], reverse=True)[:10]
        ]
    })


@router.get("/export/{report_type}")
async def export_report(
    report_type: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    导出各类报表数据 (P2-1)
    
    支持的报表类型:
    - training: 培训统计报表
    - access-sync: 门禁同步统计报表
    - access-events: 门禁事件记录报表
    - reconciliation: 对账报告
    """
    ctx = get_tenant_context()
    
    # 解析日期范围
    query_start_date = date.fromisoformat(start_date) if start_date else date.today() - timedelta(days=7)
    query_end_date = date.fromisoformat(end_date) if end_date else date.today()
    
    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    if report_type == "training":
        # 培训统计报表
        ws.title = "培训统计报表"
        
        # 表头
        headers = ["日期", "总人数", "已完成", "学习中", "未开始", "失败", "完成率(%)", "平均时长(分钟)"]
        ws.append(headers)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 获取数据
        current_date = query_start_date
        while current_date <= query_end_date:
            stmt = select(DailyTicketWorker).join(DailyTicket).where(
                DailyTicket.date == current_date,
                DailyTicketWorker.status == "ACTIVE"
            )
            stmt = TenantQueryFilter.apply(stmt, ctx)
            
            result = await db.execute(stmt)
            workers = result.scalars().all()
            
            total = len(workers)
            completed = sum(1 for w in workers if w.training_status == "COMPLETED")
            in_learning = sum(1 for w in workers if w.training_status == "IN_LEARNING")
            not_started = sum(1 for w in workers if w.training_status == "NOT_STARTED")
            failed = sum(1 for w in workers if w.training_status == "FAILED")
            completion_rate = round((completed / total * 100), 2) if total > 0 else 0
            
            # 计算平均时长
            session_stmt = select(
                func.avg(func.extract('epoch', TrainingSession.ended_at - TrainingSession.started_at))
            ).join(DailyTicketWorker).join(DailyTicket).where(
                DailyTicket.date == current_date,
                TrainingSession.status == "COMPLETED"
            )
            session_stmt = TenantQueryFilter.apply(session_stmt, ctx)
            
            session_result = await db.execute(session_stmt)
            avg_seconds = session_result.scalar() or 0
            avg_duration = round(avg_seconds / 60, 1) if avg_seconds else 0
            
            ws.append([
                current_date.strftime("%Y-%m-%d"),
                total,
                completed,
                in_learning,
                not_started,
                failed,
                completion_rate,
                avg_duration
            ])
            
            current_date += timedelta(days=1)
        
        # 调整列宽
        column_widths = [12, 10, 10, 10, 10, 10, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    elif report_type == "access-sync":
        # 门禁同步统计报表
        ws.title = "门禁同步统计报表"
        
        # 表头
        headers = ["日期", "总数", "已同步", "待同步", "失败", "已撤销", "同步率(%)", "平均时长(秒)"]
        ws.append(headers)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 获取数据
        current_date = query_start_date
        while current_date <= query_end_date:
            start_datetime = datetime.combine(current_date, datetime.min.time())
            end_datetime = datetime.combine(current_date, datetime.max.time())
            
            stmt = select(AccessGrant).where(
                AccessGrant.created_at >= start_datetime,
                AccessGrant.created_at <= end_datetime
            )
            stmt = TenantQueryFilter.apply(stmt, ctx)
            
            result = await db.execute(stmt)
            grants = result.scalars().all()
            
            total = len(grants)
            synced = sum(1 for g in grants if g.status == "SYNCED")
            pending = sum(1 for g in grants if g.status == "PENDING_SYNC")
            failed = sum(1 for g in grants if g.status == "SYNC_FAILED")
            revoked = sum(1 for g in grants if g.status == "REVOKED")
            sync_rate = round((synced / total * 100), 2) if total > 0 else 0
            
            # 计算平均同步时长
            synced_grants = [g for g in grants if g.status == "SYNCED" and g.last_sync_at]
            avg_sync_duration = 0
            if synced_grants:
                total_sync_time = sum(
                    (g.last_sync_at - g.created_at).total_seconds() 
                    for g in synced_grants 
                    if g.last_sync_at and g.created_at
                )
                avg_sync_duration = round(total_sync_time / len(synced_grants), 1)
            
            ws.append([
                current_date.strftime("%Y-%m-%d"),
                total,
                synced,
                pending,
                failed,
                revoked,
                sync_rate,
                avg_sync_duration
            ])
            
            current_date += timedelta(days=1)
        
        # 调整列宽
        column_widths = [12, 10, 10, 10, 10, 10, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    elif report_type == "access-events":
        # 门禁事件记录报表
        ws.title = "门禁事件记录"
        
        # 表头
        headers = ["时间", "工人姓名", "设备ID", "方向", "结果", "原因代码", "原因说明"]
        ws.append(headers)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 获取数据
        start_datetime = datetime.combine(query_start_date, datetime.min.time())
        end_datetime = datetime.combine(query_end_date, datetime.max.time())
        
        stmt = select(AccessEvent).where(
            AccessEvent.event_time >= start_datetime,
            AccessEvent.event_time <= end_datetime
        ).order_by(AccessEvent.event_time.desc()).limit(1000)  # 限制1000条
        
        stmt = TenantQueryFilter.apply(stmt, ctx)
        
        result = await db.execute(stmt)
        events = result.scalars().all()
        
        # 获取工人信息
        worker_ids = [e.worker_id for e in events if e.worker_id]
        if worker_ids:
            worker_stmt = select(Worker).where(Worker.worker_id.in_(worker_ids))
            worker_result = await db.execute(worker_stmt)
            workers = {w.worker_id: w.name for w in worker_result.scalars().all()}
        else:
            workers = {}
        
        # 填充数据
        direction_map = {"IN": "进入", "OUT": "离开"}
        result_map = {"PASS": "通过", "DENY": "拒绝"}
        
        for event in events:
            worker_name = workers.get(event.worker_id, "未知") if event.worker_id else "未知"
            ws.append([
                event.event_time.strftime("%Y-%m-%d %H:%M:%S"),
                worker_name,
                event.device_id or "",
                direction_map.get(event.direction, event.direction),
                result_map.get(event.result, event.result),
                event.reason_code or "",
                event.reason_message or ""
            ])
        
        # 调整列宽
        column_widths = [20, 15, 15, 10, 10, 15, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    elif report_type == "reconciliation":
        # 对账报告
        ws.title = "对账报告"
        
        # 表头
        headers = ["授权ID", "工人姓名", "区域名称", "状态", "重试次数", "最后同步时间", "错误信息"]
        ws.append(headers)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 获取卡住的授权
        cutoff = datetime.now() - timedelta(minutes=10)
        
        stmt = select(AccessGrant).join(DailyTicket).where(
            DailyTicket.date >= query_start_date,
            DailyTicket.date <= query_end_date,
            AccessGrant.status.in_(["PENDING_SYNC", "SYNC_FAILED"]),
            AccessGrant.created_at < cutoff
        )
        stmt = TenantQueryFilter.apply(stmt, ctx)
        
        result = await db.execute(stmt)
        grants = result.scalars().all()
        
        # 获取工人和区域信息
        worker_ids = [g.worker_id for g in grants]
        area_ids = [g.area_id for g in grants]
        
        if worker_ids:
            worker_stmt = select(Worker).where(Worker.worker_id.in_(worker_ids))
            worker_result = await db.execute(worker_stmt)
            workers = {w.worker_id: w.name for w in worker_result.scalars().all()}
        else:
            workers = {}
        
        if area_ids:
            area_stmt = select(WorkArea).where(WorkArea.area_id.in_(area_ids))
            area_result = await db.execute(area_stmt)
            areas = {a.area_id: a.name for a in area_result.scalars().all()}
        else:
            areas = {}
        
        # 填充数据
        status_map = {
            "PENDING_SYNC": "待同步",
            "SYNC_FAILED": "同步失败"
        }
        
        for grant in grants:
            worker_name = workers.get(grant.worker_id, "未知")
            area_name = areas.get(grant.area_id, "未知")
            
            ws.append([
                str(grant.grant_id),
                worker_name,
                area_name,
                status_map.get(grant.status, grant.status),
                grant.sync_attempt_count,
                grant.last_sync_at.strftime("%Y-%m-%d %H:%M:%S") if grant.last_sync_at else "",
                grant.sync_error_msg or ""
            ])
        
        # 调整列宽
        column_widths = [36, 15, 20, 12, 10, 20, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    else:
        from app.utils.response import error_response, ErrorCode
        return error_response(
            code=ErrorCode.VALIDATION_ERROR,
            message=f"不支持的报表类型: {report_type}"
        )
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    filename = f"{ws.title}_{query_start_date.strftime('%Y%m%d')}_{query_end_date.strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

